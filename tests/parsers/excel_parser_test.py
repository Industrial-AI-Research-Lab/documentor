import pytest
from pathlib import Path
from langchain_core.documents import Document
from documentor.parsers.excel_parser import ExcelBlobParser
from langchain_core.documents.base import Blob
import openpyxl
from PIL import Image
import io

def test_excel_parser_with_data(tmp_path):
    """
    Test to verify the correct operation of the ExcelBlobParser with real Excel data
    """
    file_path = tmp_path / "test.xlsx"
    
    # Create a test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    ws['B1'] = 'Data'
    ws['A2'] = '123'
    ws['B2'] = '456'
    wb.save(file_path)
    
    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )
    
    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))
    
    # Check basic expectations
    assert len(documents) > 0
    assert all(isinstance(doc, Document) for doc in documents)
    # Check the content
    assert 'Test;Data' in documents[0].page_content
    assert '123;456' in documents[0].page_content

def test_excel_parser_invalid_file():
    """
    Test scenario when an incorrect file path is specified.
    Expect an OSError when opening a non-existent file.
    """
    parser = ExcelBlobParser()
    blob = Blob(
        data=b"invalid data",
        path=Path("non_existent.xlsx")
    )
    
    with pytest.raises(Exception):
        list(parser.lazy_parse(blob))

def test_excel_parser_complex_data(tmp_path):
    """
    Test to verify the processing of complex Excel data, including:
    - Multiple sheets
    - Various data types
    - Empty cells
    - Formatted text
    """
    file_path = tmp_path / "test_complex.xlsx"
    
    # Create a test Excel file with multiple sheets
    wb = openpyxl.Workbook()
    
    # The first sheet - different data types
    ws1 = wb.active
    ws1.title = "Sheet1"
    test_data = [
        ["Text", "Number", "Empty", "Date"],
        ["Hello", 42, None, "2024-03-15"],
        ["World", 3.14, "", None],
        [None, 100, "Not Empty", "2024-03-16"]
    ]
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # The second sheet - text only
    ws2 = wb.create_sheet("Sheet2")
    text_data = [
        ["Name", "Description"],
        ["Product1", "Some description"],
        ["Product2", "Another description"]
    ]
    for row_idx, row_data in enumerate(text_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    
    # Test the parser
    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )
    
    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))
    
    # Check the results
    assert len(documents) == 2  # There should be two documents (one per sheet)
    
    # Check the first sheet
    sheet1_content = documents[0].page_content
    assert "Text;Number;Empty;Date" in sheet1_content
    assert "Hello;42" in sheet1_content
    assert "World;3.14" in sheet1_content
    
    # Check the second sheet
    sheet2_content = documents[1].page_content
    assert "Name;Description" in sheet2_content
    assert "Product1;Some description" in sheet2_content
    assert "Product2;Another description" in sheet2_content
    
    # Check the metadata
    for doc in documents:
        assert doc.metadata["source"] == str(file_path)
        assert doc.metadata["file_name"] == "test_complex.xlsx"
        assert doc.metadata["extension"] == ".xlsx"
        assert "sheet_name" in doc.metadata

def test_excel_parser_real_data(data_dir):
    """
    Test to verify the parsing of an Excel file with multiple sheets
    """
    file_path = Path(f"{data_dir}/test_excel.xlsx")
    
    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )
    
    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))
    
    # Check that the documents were obtained
    assert len(documents) > 0
    
    # Check that each document has data
    for doc in documents:
        assert isinstance(doc, Document)
        assert doc.page_content.strip()
        
        # Check metadata
        assert doc.metadata["source"] == str(file_path)
        assert doc.metadata["file_name"] == "test_excel.xlsx"
        assert doc.metadata["extension"] == ".xlsx"
        assert "sheet_name" in doc.metadata

def test_excel_parser_empty_sheet(data_dir):
    """
    Test to check the processing of an empty Excel sheet.
    The parser must skip empty sheets and not create documents for them.
    """
    file_path = Path(f"{data_dir}/test_empty.xlsx")

    # Create an Excel file with an empty sheet
    wb = openpyxl.Workbook()
    wb.save(file_path)

    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )

    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))

    # Check that no documents were obtained because the sheet is empty
    assert len(documents) == 0

def test_excel_parser_multiple_sheets(tmp_path):
    """
    Test to verify the processing of a file with multiple sheets
    """
    file_path = tmp_path / "test_multiple.xlsx"
    
    # Create an Excel file with two sheets
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1['A1'] = 'Sheet1 Data'
    
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = 'Sheet2 Data'
    
    wb.save(file_path)
    
    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )
    
    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))
    
    # Check that two documents were obtained
    assert len(documents) == 2
    assert all(isinstance(doc, Document) for doc in documents)
    assert all(doc.page_content.strip() for doc in documents)

def test_excel_parser_empty_and_data_sheets(data_dir):
    """
    Test to check the processing of an Excel file with different types of sheets:
    - empty sheets
    - sheets with data
    - sheets with spaces and None values
    - sheets with formulas
    
    Uses the real file test_excel.xlsx
    """
    file_path = Path(f"{data_dir}/test_excel.xlsx")
    
    with open(file_path, 'rb') as f:
        blob = Blob(
            data=f.read(),
            path=file_path
        )
    
    parser = ExcelBlobParser()
    documents = list(parser.lazy_parse(blob))
    
    # Group documents by content type
    data_docs = [doc for doc in documents if ";" in doc.page_content]
    formula_docs = [doc for doc in documents if "Formula:" in doc.page_content]
    
    # Check the number of documents with data (not formulas)
    assert len(data_docs) == 5, "Should have 5 sheets with data"

    
    # Check for the presence of formulas
    assert len(formula_docs) > 0, "Should have formulas"
    
    # Check the content of the documents
    doc_contents = [doc.page_content for doc in data_docs]
    
    # Check the first sheet (numeric data)
    first_sheet_values = ["1", "2", "3", "4", "5", "6", "7", "8", "blablabla", "4523452"]
    assert any(all(value in content for value in first_sheet_values) for content in doc_contents)
    
    # Check the second sheet (alphabetic data)
    second_sheet_values = ["a", "b", "c", "d", "e", "f", "453"]
    assert any(all(value in content for value in second_sheet_values) for content in doc_contents)
    
    # Check the fourth sheet (repeated data)
    assert any("abcd" in content and "1" in content for content in doc_contents)
    
    # Check the fifth sheet (sparse data)
    assert any("test" in content and "stse" in content for content in doc_contents)
    
    # Check the metadata for all documents
    for doc in documents:
        assert doc.metadata["source"] == str(file_path)
        assert doc.metadata["file_name"] == "test_excel.xlsx"
        assert doc.metadata["extension"] == ".xlsx"
        assert "sheet_name" in doc.metadata
