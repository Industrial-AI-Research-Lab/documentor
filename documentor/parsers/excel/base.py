from io import BytesIO
from itertools import chain, product
from typing import Iterator, Optional

import openpyxl
import pandas as pd
from langchain_core.documents import Document
from langchain_core.documents.base import Blob

from documentor.parsers.base import BaseBlobParser
from documentor.parsers.config import ParsingConfig, ParsingSchema
from documentor.parsers.extensions import DocExtension
from documentor.structuries.metadata import Metadata, to_document_metadata


class ExcelBlobParser(BaseBlobParser):
    """
    Parser for Excel files for LangChain.
    Creates a separate Document for each sheet with a text representation of the table
    and saves metadata from the original parser.
    """
    batch_lines = 0

    _extension = {DocExtension.xlsx, DocExtension.xls}
    _available_parsing_schemas = {ParsingSchema.pages}

    def __init__(self, config: Optional[ParsingConfig] = None, parse_images: Optional[bool] = None, **kwargs):
        """
        Initialize the ExcelBlobParser.

        Args:
            parse_images (Optional[bool]): Convenience flag to enable image extraction without
                constructing a custom ParsingConfig. If provided, overrides config.extract_images.
        """
        if config is None:
            config = ParsingConfig(
                parsing_schema=ParsingSchema.pages,
                extract_tables=True,
                extract_formulas=True,
                extract_images=bool(parse_images) if parse_images is not None else False,
            )
        else:
            # Respect explicit parse_images override if passed
            if parse_images is not None:
                config.extract_images = bool(parse_images)
        super().__init__(config=config, **kwargs)

    def _create_document(
            self,
            content: str,
            file_name: str,
            source: str,
            file_type: str,
            sheet_name: str = None
    ) -> Document:
        """
        Helper method to create a Document object.

        Args:
            content (str): Content of the document.
            file_name (str): Name of the file.
            source (str): Source path or identifier.
            file_type (str): Type of the file.
            sheet_name (str): Name of the Excel sheet. Defaults to None.

        Returns:
            Document: The created document.
        """
        # Build unified metadata via documentor.structuries.metadata
        meta = Metadata(
            name=file_name,
            extension=file_type,
            source=self._current_blob if hasattr(self, "_current_blob") else None,
            sheet_name=sheet_name if sheet_name else None,
        )
        metadata = to_document_metadata(
            meta,
            extension_key="extension",  # excel tests expect key "extension"
            name_key="file_name",
        )
        # Preserve explicit source string if provided
        if source is not None:
            metadata["source"] = source

        return Document(
            page_content=content,
            metadata=metadata
        )

    def _build_document(self, content: str, blob: Blob, sheet_name: str = None) -> Document:
        """
        Internal helper to remove repeated argument setup for _create_document.
        
        Args:
            content (str): Content of the document.
            blob (Blob): A Blob object containing the raw data to be parsed.
            sheet_name (str): Name of the Excel sheet. Defaults to None.

        Returns:
            Document: A Document object containing the parsed data.
        """
        # Save blob to use inside _create_document for Metadata.source
        self._current_blob = blob
        file_name = blob.path.name if blob.path else None
        source = str(blob.path) if blob.path else None
        file_type = blob.path.suffix if blob.path else None
        doc = self._create_document(content, file_name, source, file_type, sheet_name)
        if hasattr(self, "_current_blob"):
            delattr(self, "_current_blob")
        return doc

    def lazy_parse(self, blob: Blob, parse_images: bool = None) -> Iterator[Document]:
        """
        Creates a list of Document objects from an Excel file, one per sheet.
        Also extracts images if present.

        Args:
            blob (Blob): Blob object containing Excel file data
            parse_images (bool): Whether to parse images from the Excel file. 
                If None, uses the value from constructor. Defaults to None.
        Returns:
            Iterator[Document]: Iterator of Document objects
        Raises:
            Exception: if parsing fails
        """
        try:
            should_parse_images = self.config.extract_images if parse_images is None else parse_images

            excel_data = BytesIO(blob.data)
            wb = openpyxl.load_workbook(excel_data, data_only=True)

            # parse sheets
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else '' for cell in row]
                    if any(str(cell).strip() != '' for cell in cells):
                        row_content = cells
                        rows.append(row_content)

                if rows:
                    df = pd.DataFrame(rows)
                    csv_content = df.to_csv(sep=';', index=False, header=False)
                    yield self._build_document(csv_content, blob, sheet_name)

            # extract images
            if should_parse_images:
                img_content_iter = chain(*(
                    tuple(product(content._images, (content,)))
                    for content in wb.worksheets
                    if hasattr(content, '_images')
                ))

                for image, content in img_content_iter:
                    # Obtain raw image bytes from openpyxl Image object
                    img_bytes = None
                    if hasattr(image, '_data'):
                        try:
                            # In newer openpyxl versions _data is a callable returning bytes
                            img_bytes = image._data()  # type: ignore[attr-defined]
                        except TypeError:
                            # In some versions _data is already raw bytes
                            img_bytes = image._data  # type: ignore[attr-defined]
                    # Fallback: some environments may wrap bytes in BytesIO
                    if isinstance(img_bytes, BytesIO):
                        img_bytes = img_bytes.getvalue()
                    if not isinstance(img_bytes, (bytes, bytearray)):
                        # Skip if we couldn't obtain bytes
                        continue

                    img_name = f"{blob.path.stem}_image_{content.title}_{content._images.index(image)}.png"
                    with open(img_name, "wb") as f:
                        f.write(img_bytes)
                    self._logs.add_info(f"Image saved: {img_name}")

            # parse formulas
            excel_data2 = BytesIO(blob.data)
            wb_formulas = openpyxl.load_workbook(excel_data2, data_only=False)
            for sheet in wb_formulas.worksheets:
                self._logs.add_info(f"Parsing formulas on sheet: {sheet.title}")
                cells_iter = chain.from_iterable(
                    ((row_index, cell) for cell in row)
                    for row_index, row in enumerate(sheet.iter_rows())
                )
                for row_index, cell in cells_iter:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_content = f"Sheet: {sheet.title}, Cell: {cell.coordinate}, Formula: {cell.value}"
                        yield self._build_document(formula_content, blob, sheet.title)

        except Exception as e:
            raise Exception(f"An error occurred while parsing Excel file: {str(e)}") from e
