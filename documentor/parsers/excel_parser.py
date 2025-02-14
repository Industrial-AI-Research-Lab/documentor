from io import BytesIO
from itertools import chain, product
from typing import Iterator

import openpyxl
import pandas as pd
from langchain_core.documents import Document
from langchain_core.documents.base import Blob

from documentor.loaders.logger import LoaderLogger
from documentor.parsers.base import BaseBlobParser
from documentor.parsers.extensions import DocExtension


class ExcelBlobParser(BaseBlobParser):
    """
    Parser for Excel files for LangChain.
    Creates a separate Document for each sheet with a text representation of the table
    and saves metadata from the original parser.
    """
    batch_lines = 0

    _extension = {DocExtension.xlsx, DocExtension.xls}

    def __init__(self, parse_images: bool = False):
        """
        Initialize the TextBlobParser.

        Args:
            batch_lines (int): The number of lines which is one Document.
                0 value means that whole text blob is one Document. Value should be greater than or equal to 0.
            Defaults to 0.
        """
        self.parse_images = parse_images
        self._logs = LoaderLogger()

    def _create_document(self, content: str, file_name: str, source: str, file_type: str,
                         sheet_name: str = None) -> Document:
        """
        Helper method to create a Document object.

        Args:
            content (str): Content of the document.
            file_name (str): Name of the file.
            source (str): Source path or identifier.
            file_type (str): Type of the file.
            sheet_name (str): Name of the Excel sheet. Defaults to None.

        Returns:
            Document: The created document.
        """
        # TODO: decide which metadata should be used
        metadata = {
            "file_name": file_name,
            "source": source,
            "extension": file_type,
        }
        if sheet_name:
            metadata["sheet_name"] = sheet_name

        return Document(
            page_content=content,
            metadata=metadata
        )

    def _build_document(self, content: str, blob: Blob, sheet_name: str = None) -> Document:
        """
        Internal helper to remove repeated argument setup for _create_document.
        
        Args:
            content (str): Content of the document.
            blob (Blob): A Blob object containing the raw data to be parsed.
            sheet_name (str): Name of the Excel sheet. Defaults to None.

        Returns:
            Document: A Document object containing the parsed data.
        """
        file_name = blob.path.name if blob.path else None
        source = str(blob.path) if blob.path else None
        file_type = blob.path.suffix if blob.path else None
        return self._create_document(content, file_name, source, file_type, sheet_name)

    def lazy_parse(self, blob: Blob, parse_images: bool = None) -> Iterator[Document]:
        """
        Creates a list of Document objects from an Excel file, one per sheet.
        Also extracts images if present.

        Args:
            blob (Blob): Blob object containing Excel file data
            parse_images (bool): Whether to parse images from the Excel file. 
                If None, uses the value from constructor. Defaults to None.
        Returns:
            Iterator[Document]: Iterator of Document objects
        Raises:
            Exception: if parsing fails
        """
        try:
            # Используем значение из конструктора, если не передано явно
            should_parse_images = self.parse_images if parse_images is None else parse_images

            excel_data = BytesIO(blob.data)
            wb = openpyxl.load_workbook(excel_data, data_only=True)

            # parse sheets
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else '' for cell in row]
                    if any(str(cell).strip() != '' for cell in cells):
                        row_content = cells
                        rows.append(row_content)

                if rows:
                    df = pd.DataFrame(rows)
                    csv_content = df.to_csv(sep=';', index=False, header=False)
                    yield self._build_document(csv_content, blob, sheet_name)

            # extract images
            if should_parse_images:
                img_content_iter = chain(*(
                    tuple(product(content._images, (content,)))
                    for content in wb.worksheets
                    if hasattr(content, '_images')
                ))

                for image, content in img_content_iter:
                    if not hasattr(image, 'ref'):
                        continue
                    img_bytes = image.ref
                    if isinstance(img_bytes, BytesIO):
                        img_bytes = img_bytes.getvalue()
                    img_name = f"{blob.path.stem}_image_{content.title}_{content._images.index(image)}.png"
                    with open(img_name, "wb") as f:
                        f.write(img_bytes)
                    self._logs.add_info(f"Image saved: {img_name}")

            # parse formulas
            wb_formulas = openpyxl.load_workbook(excel_data, data_only=False)
            for sheet in wb_formulas.worksheets:
                self._logs.add_info(f"Parsing formulas on sheet: {sheet.title}")
                cells_iter = chain.from_iterable(
                    ((row_index, cell) for cell in row)
                    for row_index, row in enumerate(sheet.iter_rows())
                )
                for row_index, cell in cells_iter:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_content = f"Sheet: {sheet.title}, Cell: {cell.coordinate}, Formula: {cell.value}"
                        yield self._build_document(formula_content, blob, sheet.title)

        except Exception as e:
            raise Exception(f"An error occurred while parsing Excel file: {str(e)}") from e
