import openpyxl
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException

from documentor.structuries.parser import DocumentParser
from documentor.structuries.type_check import StaticClassMeta

from documentor.types.excel.document import SheetDocument

from langchain_core.documents import Document


class ExtensionException(Exception):
    """
    Exception for errors while parsing files.
    """
    pass


class ParserException(StaticClassMeta):
    """
    Static class with error messages for sheet parsing.
    """
    ExtensionException = "The {form} format is not suitable for processing."
    KeyException = "The sheet with the name {sheet_name} is not in the file."
    FileException = "The specified file does not exist or is unavailable."
    CellContentException = "A cell with this address does not exist in the table."


class SheetParser(DocumentParser):
    """
    Class for primary sheet document processing.
    """

    COLUMNS = ['value', 'start_content', 'relative_id', 'type', 'row', 'column',
               'length', 'vertically_merged', 'horizontally_merged', 'font_selection', 'top_border',
               'bottom_border', 'left_border', 'right_border', 'color', 'font_color', 'is_formula']

    def parse_file(self, path: str, sheet_name: str, first_cell: str | None = None,
                   last_cell: str | None = None) -> SheetDocument:
        """
        Create Document from file.

        :param path: path to file
        :type path: str
        :param sheet_name: name of table sheet
        :type sheet_name: str
        :param first_cell: the address of the first cell
        :type first_cell: str | None
        :param last_cell: the address of the last cell
        :type last_cell: str | None
        :return: SheetDocument object
        :rtype: SheetDocument
        :raises ExtensionException: if file extension is not supported
        :raises DocumentParsingException: if file structure is not valid
        :raises KeyError: if no sheet with the specified name in the file
        :raises ValueError: if cell address is incorrect
        :raises OSError: if file is not found or can't be opened
        """
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            wb_formulas = openpyxl.load_workbook(path)
            sheet = wb[sheet_name]
            sheet_formulas = wb_formulas[sheet_name]

            new_df = pd.DataFrame()
            merged_cells = sheet.merged_cells
            k = 0
            for column in sheet.iter_rows(
                    min_col=sheet[first_cell].column if first_cell else None,
                    min_row=sheet[first_cell].row if first_cell else None,
                    max_col=sheet[last_cell].column if last_cell else None,
                    max_row=sheet[last_cell].row if last_cell else None,
            ):
                for cel in column:
                    k += 1
                    value = cel.value
                    start_value = cel.value
                    a = str(cel).split('.')
                    is_horizontal = False
                    is_vertical = False
                    for merged_range in merged_cells:
                        if cel.coordinate in merged_range:
                            merged_range_coords = [cell for cell in merged_range.cells]
                            if len(set(list(cel)[0] for cel in merged_range_coords)) > 1:
                                is_vertical = True
                            if len(set(list(cel)[1] for cel in merged_range_coords)) > 1:
                                is_horizontal = True
                            first_cell = str(merged_range).split(':')[0]
                            if value is None:
                                value = sheet[first_cell].value
                                start_value = cel.value
                                k = k - 1

                    if len(a) > 1:
                        cell_data = [value, start_value, k, str(type(cel.value)).split("'")[1], int(cel.row),
                                     int(cel.column),
                                     len(str(cel.value)) if value else 0, is_vertical, is_horizontal, cel.font.bold,
                                     True if cel.border.top.style else False, True if cel.border.bottom.style else False,
                                     True if cel.border.left.style else False, True if cel.border.right.style else False,
                                     cel.fill.start_color.index, cel.font.color.value if cel.font.color else 0,
                                     True if cel.value != sheet_formulas[cel.coordinate].value else False]
                        fragment = pd.DataFrame(data=[cell_data], columns=self.COLUMNS)
                        new_df = pd.concat([new_df, fragment], ignore_index=True)

            return SheetDocument(df=new_df)
        except InvalidFileException as ife:
            raise InvalidFileException(ParserException.ExtensionException.format(form=path.split('.')[-1]))
        except KeyError as ke:
            raise KeyError(ParserException.KeyException.format(sheet_name=sheet_name))
        except FileNotFoundError as fnfe:
            raise FileNotFoundError(ParserException.FileException.format())
        except ValueError as ve:
            raise ValueError(ParserException.CellContentException.format())
        except Exception as e:
            raise Exception(f'{e}')

    def from_csv(self, path: str, sep: str | None) -> SheetDocument:
        """
        Create SheetDocument from csv file.

        :param path: path to file
        :param sep: separator for csv file
        :return: SheetDocument object
        :rtype: SheetDocument
        :raises DocumentParsingException: if csv structure is not valid
        :raises OSError: if file is not found or can't be opened
        """
        df = pd.read_csv(path, sep)
        return SheetDocument(df)

    def to_csv(self, document: SheetDocument, path: str, sep: str | None = None):
        """
        Save SheetDocument to csv file.

        :param document: SheetDocument object for storing in csv file
        :type document: SheetDocument
        :param path: path to file
        :type path: str
        :param sep: separator for csv file
        :type sep: str | None
        :raises OSError: if document can't be written to file
        """
        document.to_df().to_csv(path, sep=sep if sep else ",")


class SheetLangChainParser(DocumentParser):
    """
    Parser for Excel files for LangChain.
    Creates a separate Document for each sheet with a text representation of the table
    and saves metadata from the original parser.
    """


    def __init__(self):
        self.sheet_parser = SheetParser()

    def parse_file(self, path: str) -> list[Document]:
        """
        Creates a list of Document objects from an Excel file, one per sheet.

        Args:
            path (str): path to file
        Returns:
            list[Document]: list of Document objects
        Raises:
            ExtensionException: if file extension is not supported
            OSError: if file is not found or unavailable
        """
        try:


            wb = openpyxl.load_workbook(path, data_only=True)
            documents = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get metadata through the original parser
                try:
                    meta_document = self.sheet_parser.parse_file(path, sheet_name)
                    meta_df = meta_document.to_df()

                    
                    # Create a text representation of the table based on metadata
                    table_data = []
                    current_row = []
                    last_row = -1
                    

                    # Sort cells by rows and columns
                    sorted_cells = meta_df.sort_values(['row', 'column'])
                    

                    for _, cell in sorted_cells.iterrows():
                        if cell['row'] != last_row and last_row != -1:
                            table_data.append(current_row)
                            current_row = []
                        
                        # Fill missing columns with empty values
                        while len(current_row) < cell['column'] - 1:
                            current_row.append('')
                            

                        current_row.append(str(cell['value']) if cell['value'] is not None else '')
                        last_row = cell['row']
                    
                    if current_row:
                        table_data.append(current_row)
                    
                    # Create a CSV representation
                    df = pd.DataFrame(table_data)
                    csv_content = df.to_csv(sep=';', index=False, header=False)
                    

                    metadata = {
                        "sheet_name": sheet_name,
                        "file_path": path,
                        "sheet_index": wb.sheetnames.index(sheet_name),
                        "meta_df": meta_df
                    }
                    
                except Exception as e:
                    # If it is not possible to get metadata, create a simple representation
                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        data.append([str(cell) if cell is not None else '' for cell in row])

                    
                    df = pd.DataFrame(data)
                    csv_content = df.to_csv(sep=';', index=False, header=False)
                    
                    metadata = {
                        "sheet_name": sheet_name,
                        "file_path": path,
                        "sheet_index": wb.sheetnames.index(sheet_name),
                        "parse_error": str(e)
                    }

                document = Document(
                    page_content=csv_content,
                    metadata=metadata
                )
                documents.append(document)

            return documents

        except InvalidFileException:
            raise InvalidFileException(ParserException.ExtensionException.format(form=path.split('.')[-1]))
        except FileNotFoundError:
            raise FileNotFoundError(ParserException.FileException)
        except Exception as e:
            raise Exception(f'Error parsing file: {e}')